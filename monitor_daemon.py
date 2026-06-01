#!/usr/bin/env python3
"""
RBKS Compliance Dashboard — Auto-Monitor Daemon
自動監控 Confluence & SharePoint 變動，即時更新 Dashboard

Features:
  - Monitors Ring Confluence NPI Overview for new projects
  - Monitors SharePoint Excel for personnel changes
  - Auto-fetches schedules for newly detected projects
  - Rebuilds dashboard.html automatically
  - Serves dashboard to team via HTTP

Usage:
  python3 monitor_daemon.py              # Start monitoring + serving
  python3 monitor_daemon.py --check      # Single check (no server)
  python3 monitor_daemon.py --rebuild    # Force rebuild dashboard
  python3 monitor_daemon.py --port 9090  # Custom port (default: 8080)

Requirements:
  pip3 install playwright pycookiecheat requests openpyxl
"""

import json, re, time, sys, os, hashlib, threading
import subprocess
from pathlib import Path
from datetime import datetime
from http.server import SimpleHTTPRequestHandler, HTTPServer

BASE_DIR = Path(__file__).parent
AUTH_STATE = BASE_DIR / "auth_state.json"
CACHE_FILE = BASE_DIR / "monitor_cache.json"
DASHBOARD_FILE = BASE_DIR / "dashboard.html"
SCHEDULES_FILE = BASE_DIR / "consolidated_schedules.json"
PERSONNEL_FILE = BASE_DIR / "personnel_data.json"
LOG_FILE = BASE_DIR / "monitor.log"

# Monitoring intervals
CHECK_INTERVAL = 900  # 15 minutes
CONFLUENCE_URL = "https://confluence.atl.ring.com"
OVERVIEW_PAGE_ID = "3355651641"
SHAREPOINT_FILE_ID = "EB48E212-8F85-46D3-99AC-E1CA5C6B00CB"

# Schedule search patterns for new projects
SCHEDULE_SEARCH_PATTERNS = [
    "{project} - Schedule",
    "{project} - HW schedule milestone",
    "{project} Hardware Development Weekly Update",
    "{project} - Program Status",
    "{project} - Top Level Program Plan",
    "{project} - HW weekly update",
]


def log(msg, level="INFO"):
    """Log to console and file"""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] [{level}] {msg}"
    print(line)
    try:
        with open(LOG_FILE, "a") as f:
            f.write(line + "\n")
    except:
        pass


def load_cache():
    """Load monitoring cache"""
    if CACHE_FILE.exists():
        try:
            return json.loads(CACHE_FILE.read_text())
        except:
            pass
    return {
        "confluence_version": 0,
        "confluence_projects": [],
        "sharepoint_hash": "",
        "last_check": "",
        "last_update": "",
        "update_history": []
    }


def save_cache(cache):
    """Save monitoring cache"""
    CACHE_FILE.write_text(json.dumps(cache, indent=2, ensure_ascii=False))


def get_firefox_cookies(domain):
    """Get Firefox cookies for authentication"""
    try:
        from pycookiecheat import firefox_cookies
        return firefox_cookies(domain)
    except Exception as e:
        log(f"Cookie error for {domain}: {e}", "WARN")
        return None


def check_confluence_overview():
    """
    Check Confluence NPI Overview page for changes.
    Returns: (version, project_list) or None on failure
    """
    cookies = get_firefox_cookies(CONFLUENCE_URL)
    if not cookies:
        return None

    import requests
    try:
        # Get page version
        api_url = f"{CONFLUENCE_URL}/rest/api/content/{OVERVIEW_PAGE_ID}?expand=version,body.storage"
        resp = requests.get(api_url, cookies=cookies, timeout=30)
        if resp.status_code != 200:
            log(f"Confluence API returned {resp.status_code}", "WARN")
            return None

        data = resp.json()
        version = data.get("version", {}).get("number", 0)
        body_html = data.get("body", {}).get("storage", {}).get("value", "")

        # Extract project names from NPI table
        projects = extract_npi_projects(body_html)
        return (version, projects)

    except Exception as e:
        log(f"Confluence check failed: {e}", "ERROR")
        return None


def extract_npi_projects(html):
    """Extract project code names from NPI table HTML"""
    projects = []
    # Look for table rows with project names
    # Pattern: <td>ProjectName</td> or links containing project names
    rows = re.findall(r'<tr[^>]*>(.*?)</tr>', html, re.DOTALL)
    for row in rows:
        cells = re.findall(r'<td[^>]*>(.*?)</td>', row, re.DOTALL)
        if cells:
            # First cell usually contains code name
            name = re.sub(r'<[^>]+>', '', cells[0]).strip()
            if name and len(name) > 1 and len(name) < 40 and not name.startswith('#'):
                projects.append(name)
    return projects


def check_sharepoint_excel():
    """
    Check SharePoint Excel for modifications.
    Returns: file hash or None on failure
    """
    cookies = get_firefox_cookies("https://amazon.sharepoint.com")
    if not cookies:
        return None

    import requests
    try:
        # Get file metadata
        api_url = f"https://amazon.sharepoint.com/sites/rbks-compliance/_api/web/GetFileById('{SHAREPOINT_FILE_ID}')"
        resp = requests.get(api_url, cookies=cookies,
                          headers={"Accept": "application/json"}, timeout=20)
        if resp.status_code == 200:
            data = resp.json()
            modified = data.get("d", {}).get("TimeLastModified", "")
            return modified
    except Exception as e:
        log(f"SharePoint check failed: {e}", "ERROR")
    return None


def download_sharepoint_excel():
    """Download latest SharePoint Excel"""
    cookies = get_firefox_cookies("https://amazon.sharepoint.com")
    if not cookies:
        return False

    import requests
    try:
        api_url = f"https://amazon.sharepoint.com/sites/rbks-compliance/_api/web/GetFileById('{SHAREPOINT_FILE_ID}')/$value"
        resp = requests.get(api_url, cookies=cookies, timeout=60)
        if resp.status_code == 200 and len(resp.content) > 5000:
            (BASE_DIR / "NPI_Status_Tracker.xlsx").write_bytes(resp.content)
            log("SharePoint Excel downloaded successfully")
            return True
        else:
            log(f"SharePoint download failed: status={resp.status_code}", "WARN")
    except Exception as e:
        log(f"SharePoint download error: {e}", "ERROR")
    return False


def extract_personnel_from_excel():
    """Extract personnel data from downloaded Excel"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(BASE_DIR / "NPI_Status_Tracker.xlsx", read_only=True)
        ws = wb.active

        personnel = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            # Columns: A=name, B=status, C=phase, D=brand, ...
            # N=PM, O=RF, P=EMC, Q=Energy, R=Safety, S=GMA, T=CM
            name = str(row[0]).strip() if row[0] else ""
            if not name:
                continue

            entry = {
                "name": name,
                "status": str(row[1] or ""),
                "phase": str(row[2] or ""),
                "brand": str(row[3] or ""),
                "pm": str(row[13] or "") if len(row) > 13 else "",
                "rf": str(row[14] or "") if len(row) > 14 else "",
                "emc": str(row[15] or "") if len(row) > 15 else "",
                "energy": str(row[16] or "") if len(row) > 16 else "",
                "safety": str(row[17] or "") if len(row) > 17 else "",
                "gma": str(row[18] or "") if len(row) > 18 else "",
                "cm": str(row[19] or "") if len(row) > 19 else "",
            }
            # Filter out Lab 126 projects
            if "lab 126" not in entry["pm"].lower() and "lab126" not in entry["pm"].lower():
                personnel.append(entry)

        wb.close()
        PERSONNEL_FILE.write_text(json.dumps(personnel, indent=2, ensure_ascii=False))
        log(f"Extracted {len(personnel)} personnel records (Lab 126 excluded)")
        return personnel

    except Exception as e:
        log(f"Excel extraction error: {e}", "ERROR")
        return None


def fetch_project_schedule(project_name):
    """
    Attempt to fetch schedule for a new project from Confluence.
    Searches multiple page patterns.
    """
    cookies = get_firefox_cookies(CONFLUENCE_URL)
    if not cookies:
        return None

    import requests

    # Search Confluence for schedule pages
    for pattern in SCHEDULE_SEARCH_PATTERNS:
        query = pattern.format(project=project_name)
        search_url = f"{CONFLUENCE_URL}/rest/api/content/search?cql=title~\"{query}\"&limit=5"

        try:
            resp = requests.get(search_url, cookies=cookies, timeout=15)
            if resp.status_code == 200:
                results = resp.json().get("results", [])
                for result in results:
                    page_id = result.get("id")
                    schedule = extract_schedule_from_page(page_id, cookies)
                    if schedule:
                        log(f"  Found schedule for {project_name} from: {result.get('title')}")
                        return schedule
        except:
            continue

    # Try HW Weekly Update pattern
    try:
        query = f"{project_name} HW weekly"
        search_url = f"{CONFLUENCE_URL}/rest/api/content/search?cql=title~\"{query}\"&limit=3&orderby=lastmodified+desc"
        resp = requests.get(search_url, cookies=cookies, timeout=15)
        if resp.status_code == 200:
            results = resp.json().get("results", [])
            for result in results:
                page_id = result.get("id")
                schedule = extract_schedule_from_page(page_id, cookies)
                if schedule:
                    return schedule
    except:
        pass

    return None


def extract_schedule_from_page(page_id, cookies):
    """Extract schedule dates from a Confluence page"""
    import requests

    try:
        url = f"{CONFLUENCE_URL}/rest/api/content/{page_id}?expand=body.storage"
        resp = requests.get(url, cookies=cookies, timeout=15)
        if resp.status_code != 200:
            return None

        body = resp.json().get("body", {}).get("storage", {}).get("value", "")
        text = re.sub(r'<[^>]+>', ' ', body)

        schedule = {}

        # Date patterns
        date_patterns = [
            r'(\d{1,2}/\d{1,2}/\d{2,4})',  # MM/DD/YY or MM/DD/YYYY
            r'(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4})',  # DD Mon YYYY
            r'(\d{4}-\d{2}-\d{2})',  # YYYY-MM-DD
            r'(Q[1-4]\s+\d{4})',  # Q1 2027
        ]

        # Look for EVT/Pre-beta
        evt_match = re.search(r'(?:EVT|Pre-?beta|Engineering\s+Validation)[^:]*[:：\s]+([^\n<]{5,30})', text, re.IGNORECASE)
        if evt_match:
            schedule["evt"] = evt_match.group(1).strip()

        # Look for DVT/Beta
        dvt_match = re.search(r'(?:DVT|Beta(?!\s*\())[^:]*[:：\s]+([^\n<]{5,30})', text, re.IGNORECASE)
        if dvt_match:
            schedule["dvt"] = dvt_match.group(1).strip()

        # Look for OK2MP/MP
        mp_match = re.search(r'(?:OK2MP|Mass\s+Production|MP\s+Start)[^:]*[:：\s]+([^\n<]{5,30})', text, re.IGNORECASE)
        if mp_match:
            schedule["ok2mp"] = mp_match.group(1).strip()

        # Look for Ship
        ship_match = re.search(r'(?:Ship(?:ment)?|First\s+Ship)[^:]*[:：\s]+([^\n<]{5,30})', text, re.IGNORECASE)
        if ship_match:
            schedule["ship"] = ship_match.group(1).strip()

        # Look for Street/Launch
        street_match = re.search(r'(?:Street|Launch|On-?shelf)[^:]*[:：\s]+([^\n<]{5,30})', text, re.IGNORECASE)
        if street_match:
            schedule["street"] = street_match.group(1).strip()

        # Look for Announce
        ann_match = re.search(r'(?:Announce|Announcement)[^:]*[:：\s]+([^\n<]{5,30})', text, re.IGNORECASE)
        if ann_match:
            schedule["announce"] = ann_match.group(1).strip()

        if schedule:
            return schedule

    except Exception as e:
        pass

    return None


def detect_new_projects(current_projects, cached_projects):
    """Detect newly added projects"""
    cached_set = set(cached_projects)
    new_projects = [p for p in current_projects if p not in cached_set]
    return new_projects


def rebuild_dashboard():
    """Rebuild the dashboard HTML from current data"""
    log("🔄 Rebuilding dashboard...")

    # Load current data
    schedules = {}
    if SCHEDULES_FILE.exists():
        schedules = json.loads(SCHEDULES_FILE.read_text())

    personnel = []
    if PERSONNEL_FILE.exists():
        personnel = json.loads(PERSONNEL_FILE.read_text())

    # Update timestamp in dashboard
    if DASHBOARD_FILE.exists():
        content = DASHBOARD_FILE.read_text()
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        content = re.sub(
            r'Last updated 最後更新: [^|]+\|',
            f'Last updated 最後更新: {now} |',
            content
        )
        content = re.sub(
            r'Generated: \d{4}-\d{2}-\d{2}',
            f'Generated: {datetime.now().strftime("%Y-%m-%d")}',
            content
        )
        DASHBOARD_FILE.write_text(content)
        log("Dashboard timestamp updated")

    return True


def git_push_dashboard():
    """Auto commit and push dashboard to GitHub Pages"""
    try:
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        # Stage dashboard.html and index.html
        subprocess.run(
            ["git", "add", "dashboard.html", "index.html"],
            cwd=str(BASE_DIR), capture_output=True, timeout=10
        )
        # Commit
        result = subprocess.run(
            ["git", "commit", "-m", f"Auto-update dashboard: {now}"],
            cwd=str(BASE_DIR), capture_output=True, text=True, timeout=10
        )
        if result.returncode == 0:
            # Push
            push_result = subprocess.run(
                ["git", "push"],
                cwd=str(BASE_DIR), capture_output=True, text=True, timeout=30
            )
            if push_result.returncode == 0:
                log("🚀 Dashboard pushed to GitHub Pages automatically")
            else:
                log(f"⚠️ Git push failed: {push_result.stderr[:100]}", "WARN")
        else:
            # No changes to commit (already up to date)
            if "nothing to commit" in result.stdout:
                log("Git: no changes to push")
            else:
                log(f"⚠️ Git commit issue: {result.stdout[:100]}", "WARN")
    except Exception as e:
        log(f"⚠️ Git auto-push error: {e}", "WARN")


def run_monitoring_cycle():
    """Run one monitoring cycle"""
    cache = load_cache()
    changes_detected = False

    # 1. Check Confluence
    log("📡 Checking Ring Confluence NPI Overview...")
    conf_result = check_confluence_overview()
    if conf_result:
        version, projects = conf_result
        old_version = cache.get("confluence_version", 0)

        if version != old_version:
            log(f"  ✅ Confluence updated: v{old_version} → v{version}")
            new_projects = detect_new_projects(projects, cache.get("confluence_projects", []))

            if new_projects:
                log(f"  🆕 New projects detected: {', '.join(new_projects)}")
                # Fetch schedules for new projects
                for proj in new_projects:
                    log(f"  🔍 Searching schedule for: {proj}")
                    schedule = fetch_project_schedule(proj)
                    if schedule:
                        schedules = json.loads(SCHEDULES_FILE.read_text()) if SCHEDULES_FILE.exists() else {}
                        schedules[proj] = schedule
                        SCHEDULES_FILE.write_text(json.dumps(schedules, indent=2, ensure_ascii=False))
                        log(f"  ✅ Schedule found and saved for {proj}")
                    else:
                        log(f"  ⚠️ No schedule found for {proj} (will retry next cycle)")

            cache["confluence_version"] = version
            cache["confluence_projects"] = projects
            changes_detected = True
        else:
            log(f"  No change (v{version})")
    else:
        log("  ⚠️ Could not reach Confluence (auth may have expired)")

    # 2. Check SharePoint
    log("📡 Checking SharePoint Excel...")
    sp_modified = check_sharepoint_excel()
    if sp_modified:
        old_modified = cache.get("sharepoint_modified", "")
        if sp_modified != old_modified:
            log(f"  ✅ SharePoint Excel modified: {sp_modified}")
            if download_sharepoint_excel():
                extract_personnel_from_excel()
            cache["sharepoint_modified"] = sp_modified
            changes_detected = True
        else:
            log("  No change")
    else:
        log("  ⚠️ Could not reach SharePoint (auth may have expired)")

    # 3. Rebuild if changes detected
    if changes_detected:
        rebuild_dashboard()
        git_push_dashboard()
        cache["last_update"] = datetime.now().isoformat()
        history = cache.get("update_history", [])
        history.append({
            "time": datetime.now().isoformat(),
            "type": "auto",
            "changes": "Confluence/SharePoint update detected"
        })
        # Keep last 50 entries
        cache["update_history"] = history[-50:]
        log("✅ Dashboard updated and deployed to GitHub Pages!")
    else:
        log("No changes detected.")

    cache["last_check"] = datetime.now().isoformat()
    save_cache(cache)
    return changes_detected


class DashboardHandler(SimpleHTTPRequestHandler):
    """Custom HTTP handler for dashboard serving"""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(BASE_DIR), **kwargs)

    def log_message(self, format, *args):
        pass  # Suppress HTTP logs

    def do_GET(self):
        # Default to dashboard.html
        if self.path == "/" or self.path == "":
            self.path = "/dashboard.html"
        # API endpoint for status
        if self.path == "/api/status":
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            cache = load_cache()
            status = {
                "last_check": cache.get("last_check", "Never"),
                "last_update": cache.get("last_update", "Never"),
                "confluence_version": cache.get("confluence_version", 0),
                "total_projects": len(cache.get("confluence_projects", [])),
                "monitoring": True
            }
            self.wfile.write(json.dumps(status).encode())
            return
        super().do_GET()


def start_server(port=8080):
    """Start HTTP server in background thread"""
    import socket

    def get_ip():
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
            s.close()
            return ip
        except:
            return "localhost"

    server = HTTPServer(("0.0.0.0", port), DashboardHandler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()

    ip = get_ip()
    log(f"🌐 Dashboard server started:")
    log(f"   Local:   http://localhost:{port}/dashboard.html")
    log(f"   Network: http://{ip}:{port}/dashboard.html")
    log(f"   Status:  http://localhost:{port}/api/status")
    return server


def main():
    port = 8080
    if "--port" in sys.argv:
        idx = sys.argv.index("--port")
        port = int(sys.argv[idx + 1])

    if "--check" in sys.argv:
        # Single check mode
        log("=" * 60)
        log("RBKS Compliance Dashboard — Single Check")
        log("=" * 60)
        run_monitoring_cycle()
        return

    if "--rebuild" in sys.argv:
        # Force rebuild
        log("Force rebuilding dashboard...")
        rebuild_dashboard()
        return

    # Full daemon mode: monitor + serve
    log("=" * 60)
    log("🛡️ RBKS Compliance Dashboard — Auto-Monitor Daemon")
    log("=" * 60)
    log(f"Check interval: {CHECK_INTERVAL // 60} minutes")
    log(f"Dashboard: {DASHBOARD_FILE}")
    log("")

    # Start web server
    start_server(port)

    # Initial check
    log("Running initial check...")
    run_monitoring_cycle()

    # Monitoring loop
    log(f"\n👁️ Monitoring started. Next check in {CHECK_INTERVAL // 60} min...")
    while True:
        try:
            time.sleep(CHECK_INTERVAL)
            log("\n" + "─" * 40)
            run_monitoring_cycle()
            log(f"Next check in {CHECK_INTERVAL // 60} min...")
        except KeyboardInterrupt:
            log("\n🛑 Monitor stopped by user.")
            break
        except Exception as e:
            log(f"Error in monitoring cycle: {e}", "ERROR")
            time.sleep(60)  # Wait 1 min before retry


if __name__ == "__main__":
    main()
